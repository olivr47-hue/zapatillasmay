# -*- coding: utf-8 -*-
"""
exportar_mercadolibre.py
========================
Lee todos los productos activos del ERP (Supabase) y genera la plantilla
de publicación masiva de MercadoLibre lista para subir.

Uso:
    python exportar_mercadolibre.py

Requiere:
    pip install openpyxl python-dotenv

Salida:
    ML_Zapatillas_May_<fecha>.xlsx  (en la misma carpeta donde está este script)

Notas:
    - Las fotos son URLs de Cloudinary. Antes de subir a ML tendrás que
      cargarlas en el Gestor de fotos de ML y reemplazar las URLs.
    - El precio usa precio_menudeo + envío gratis (tú absorbes el costo de envío).
    - La guía de tallas por defecto es la más amplia que tengas en tu cuenta (código 1339758).
      Verifica que sea la correcta antes de publicar.
"""

import os, json, sys, urllib.request
from datetime import datetime
from copy import copy
from dotenv import load_dotenv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: Instala openpyxl primero:  pip install openpyxl")
    sys.exit(1)

# ─── Configuración ────────────────────────────────────────────────────────────

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://zulrlnivrhvwfexqktes.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")

# Ruta al template original de ML (para preservar formato y metadatos)
TEMPLATE_DEFAULT = os.path.join(os.path.expanduser("~"), "Downloads", "Publicar-05-27-17_52_30.xlsx")

# Código de la guía de tallas (la más amplia: cubre 22-27 MX, mujeres)
GUIA_MUJER   = "1339758"
GUIA_NINA    = "1339758"   # ajusta si tienes una guía específica para niñas

# Campos fijos
MARCA        = "May"
CONDICION    = "Nuevo"
TIPO_PUB     = "Clásica"
FORMA_ENVIO  = "Mercado Envíos"
COSTO_ENVIO  = "Ofreces envío gratis"
RETIRO       = "Acepto"
MONEDA       = "$"
COD_UNIV     = "El producto no tiene código registrado"

# ─── Normalización de colores ERP -> ML ───────────────────────────────────────

COLOR_ML = {
    # Negros
    "NEGRO": "Negro", "NEGRO CHAROL": "Negro", "NEGRO NAPA": "Negro",
    "NEGRO MATE": "Negro", "NEGRO BRILLANTE": "Negro",
    # Blancos
    "BLANCO": "Blanco", "BLANCO CHAROL": "Blanco",
    # Beiges / Cremas
    "BEIGE": "Beige", "AVENA": "Beige", "ICE": "Beige", "CREMA": "Crema",
    "LINO": "Beige", "HUESO": "Crema",
    # Nude
    "NUDE": "Nude", "NUDE CHAROL": "Nude", "NUDE NAPA": "Nude",
    "NUDE MATE": "Nude",
    # Marrones / Cafés
    "CAFE": "Marrón", "CAFÉ": "Marrón", "CAFE CHAROL": "Marrón",
    "CHOCOLATE": "Chocolate", "CAFE SERPIENTE": "Marrón oscuro",
    "CAFE DURAZNO": "Suela", "CAPUCHINO": "Marrón claro",
    "COGNAC": "Marrón claro", "COGÑAC": "Marrón claro", "Cognac": "Marrón claro",
    "CAMEL": "Marrón claro", "CAREY": "Marrón", "CAREY CHAROL": "Marrón",
    "Carey charol": "Marrón", "CAFE CHAROL": "Marrón",
    # Latte / Arena
    "LATTE": "Suela", "LATE": "Suela", "ARENILLA": "Suela",
    "LATTE/ NEGRO": "Negro", "LATTE//NEGRO": "Negro",
    "DESIERTO": "Ocre",
    # Colores adicionales del ERP
    "MIGA": "Marrón claro", "MOKA": "Marrón oscuro", "MOCHA": "Marrón oscuro",
    "MAQUILLAJE": "Nude", "MAQUILLAJE CHAROL": "Nude", "MAQUILLAJE NAPA": "Nude",
    "MAQUILAJE CHAROL": "Nude", "MAQUILAJE": "Nude",
    "NATURAL": "Nude", "NEUTRO": "Beige", "SANDALO": "Suela",
    "SEDA": "Crema", "TAN": "Suela", "TANG": "Suela",
    "VACA": "Marrón", "YESO": "Blanco", "MADERA": "Marrón claro",
    "DORADO OSCURO": "Dorado oscuro",
    # Dorados / Metálicos
    "ORO": "Dorado", "ORO METALICO": "Dorado", "DORADO": "Dorado",
    "INOX": "Dorado", "Inox": "Dorado", "ORO ROSA": "Dorado oscuro",
    # Plateados
    "PLATEADO": "Plateado", "PLATA": "Plateado",
    # Rojos
    "ROJO": "Rojo", "CORAL": "Coral", "SALMON": "Naranja claro",
    # Rosas
    "ROSA": "Rosa", "ROSA CHAROL": "Rosa", "ROSA CLARO": "Rosa claro",
    "FUCSIA": "Fucsia", "FUSHA": "Fucsia",
    # Vinos / Bordeaux
    "VINO": "Bordo", "BORDO": "Bordo", "BURDEOS": "Bordo",
    # Azules
    "AZUL MARINO": "Azul marino", "AUL MARINO": "Azul marino",
    "Azul marino charol": "Azul marino",
    "AZUL": "Azul", "AZUL MEZCLILLA": "Azul", "AZUL CLARO": "Azul claro",
    # Verdes
    "VERDE": "Verde", "VERDE MENTA": "Verde claro", "VERDE MUSGO": "Verde musgo",
    # Morados
    "MORADO": "Violeta", "LILA": "Lila", "LAVANDA": "Lavanda",
    # Amarillos / Naranja
    "AMARILLO": "Amarillo", "AMARILLO CHAROL": "Amarillo",
    "NARANJA": "Naranja",
    # Grises
    "GRIS": "Gris", "GRIS CLARO": "Gris", "GRIS OSCURO": "Gris oscuro",
    "TAUPE": "Gris",
    # Multicolor
    "MULTICOLOR": "Escribe o elige un valor",
}

def normalizar_color(color_erp):
    """Normaliza el color del ERP al nombre que acepta ML."""
    if not color_erp:
        return "Escribe o elige un valor"
    c = color_erp.strip()
    # Búsqueda exacta
    if c in COLOR_ML:
        return COLOR_ML[c]
    # Búsqueda case-insensitive
    cu = c.upper()
    for k, v in COLOR_ML.items():
        if k.upper() == cu:
            return v
    # Búsqueda parcial: si el color contiene una palabra clave
    for k, v in COLOR_ML.items():
        if k.upper() in cu or cu in k.upper():
            return v
    # Fallback: devolver tal cual (el usuario puede corregir)
    return c.capitalize()

# ─── Títulos descriptivos ─────────────────────────────────────────────────────

TIPO_TACON_LABEL = {
    "aguja":     "Tacón aguja",
    "plataforma":"Plataforma",
    "cuña":      "Cuña",
    "bajo":      "Tacón bajo",
    "plano":     "Zapato plano",
    "sandalia":  "Sandalia tacón",
    "palo":      "Tacón palo",
    "ancho":     "Tacón ancho",
    "cono":      "Tacón cono",
    "cuadrado":  "Tacón cuadrado",
    "embudo":    "Tacón embudo",
}

CAT_LABEL = {
    "tacones":    "Tacón",
    "sandalias":  "Sandalia",
    "botas":      "Bota",
    "botines":    "Botín",
    "flats":      "Zapato plano",
    "plataformas":"Plataforma",
    "nina":       "Zapato",
    "accesorios": "Accesorio",
}

MAT_TITULO = {
    "SINTETICO": "sintético", "SINTÉTICO": "sintético",
    "CHAROL": "charol", "CUERO": "cuero", "PIEL": "piel",
    "GAMUZA": "gamuza", "LONA": "lona", "TELA": "tela",
}

def gen_titulo(prod):
    cat    = (prod.get("categoria") or "").lower()
    tipo   = (prod.get("tipo_tacon") or "").lower()
    altura = prod.get("altura_tacon")
    mat_raw = (prod.get("material") or "SINTETICO").strip().upper()
    mat_l  = MAT_TITULO.get(mat_raw, "sintético")
    genero = "niña" if cat == "nina" else "mujer"

    tipo_label = TIPO_TACON_LABEL.get(tipo, CAT_LABEL.get(cat, "Zapato"))

    if altura and float(altura) > 0:
        h = int(float(altura)) if float(altura) == int(float(altura)) else float(altura)
        titulo = f"{tipo_label} {mat_l} {h}cm May {genero}"
    else:
        titulo = f"{tipo_label} {mat_l} May {genero}"

    return titulo[:60].strip()

# ─── Mapeo de categorías a hojas ML ──────────────────────────────────────────

# Nombre exacto de la hoja en el template
CAT_TO_SHEET = {
    "tacones":    "Zapatillas y Tacones",
    "botas":      "Zapatillas y Tacones",
    "botines":    "Zapatillas y Tacones",
    "plataformas":"Zapatillas y Tacones",
    "nina":       "Zapatillas y Tacones",
    "sandalias":  "Sandalias y Chanclas",
    "flats":      "Mocasines y Oxfords",
    "accesorios": None,  # se omite
}

# Tipo de calzado (solo para Zapatillas y Tacones)
CAT_TO_TIPO_CALZADO = {
    "plataformas": "Plataformas",
    "tacones":     "Tacos",
    "botas":       "Tacos",
    "botines":     "Tacos",
    "nina":        "Tacos",
}

TIPO_TACON_ML = {
    "aguja": "Aguja", "palo": "Palo", "bajo": "Bajo", "cuña": "Cuña",
    "ancho": "Ancho", "cono": "Cono", "cuadrado": "Cuadrado",
    "embudo": "Embudo", "plataforma": None,
}

MATERIAL_ML = {
    "SINTETICO": "Cuero sintético", "SINTÉTICO": "Cuero sintético",
    "CHAROL": "Charol", "CUERO": "Piel", "GAMUZA": "Gamuza",
    "LONA": "Lona", "PIEL": "Piel",
}

def normalizar_material(mat):
    if not mat:
        return None
    m = mat.strip().upper()
    return MATERIAL_ML.get(m, "Cuero sintético")

# ─── Supabase helpers ─────────────────────────────────────────────────────────

_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
}

def _get(path, offset=0, limit=1000):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {**_HEADERS, "Range-Unit": "items", "Range": f"{offset}-{offset+limit-1}"}
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        if e.code == 416:
            return []
        raise

def get_all(path):
    todos, offset = [], 0
    while True:
        chunk = _get(path, offset)
        if not isinstance(chunk, list) or not chunk:
            break
        todos.extend(chunk)
        if len(chunk) < 1000:
            break
        offset += 1000
    return todos

# ─── Carga de datos ───────────────────────────────────────────────────────────

def cargar_datos():
    print("[1/3] Cargando productos…")
    productos = get_all(
        "productos?activo=eq.true"
        "&select=id,nombre,sku_interno,categoria,tipo_tacon,altura_tacon,"
        "precio_menudeo,material,material_suela,descripcion"
    )
    prod_map = {p["id"]: p for p in productos}
    print(f"   -> {len(productos)} productos activos")

    print("[2/3] Cargando variantes…")
    variantes = get_all(
        "variantes?activa=eq.true"
        "&select=id,producto_id,color,talla,sku,foto_url,imagenes"
    )
    print(f"   -> {len(variantes)} variantes activas")

    print("[3/3] Cargando inventario…")
    inventario_raw = get_all(
        "inventario?select=variante_id,cantidad"
    )
    # Sumar stock por variante (puede haber varias sucursales)
    stock_map = {}
    for row in inventario_raw:
        vid = row.get("variante_id")
        qty = row.get("cantidad") or 0
        stock_map[vid] = stock_map.get(vid, 0) + qty
    print(f"   -> {len(stock_map)} registros de inventario")

    return prod_map, variantes, stock_map

# ─── Construcción de filas ────────────────────────────────────────────────────

def construir_filas(prod_map, variantes, stock_map):
    """Devuelve un dict {nombre_hoja: [filas]}."""
    sheets = {
        "Zapatillas y Tacones": [],
        "Sandalias y Chanclas": [],
        "Mocasines y Oxfords":  [],
    }

    # Agrupar variantes por producto para obtener el título una sola vez
    por_producto = {}
    for v in variantes:
        pid = v.get("producto_id")
        if pid not in por_producto:
            por_producto[pid] = []
        por_producto[pid].append(v)

    for pid, vars_prod in por_producto.items():
        prod = prod_map.get(pid)
        if not prod:
            continue

        cat    = (prod.get("categoria") or "").lower()
        sheet  = CAT_TO_SHEET.get(cat)
        if not sheet:
            continue

        titulo   = gen_titulo(prod)
        modelo   = (prod.get("nombre") or "").strip()
        precio   = prod.get("precio_menudeo") or 0
        mat_cal  = normalizar_material(prod.get("material"))
        mat_suel = normalizar_material(prod.get("material_suela"))
        desc     = (prod.get("descripcion") or "").strip()[:500]
        genero   = "Niñas" if cat == "nina" else "Mujer"
        guia     = GUIA_NINA if cat == "nina" else GUIA_MUJER

        tipo_cal = CAT_TO_TIPO_CALZADO.get(cat, "")
        tipo_tac = TIPO_TACON_ML.get((prod.get("tipo_tacon") or "").lower(), None)

        for v in vars_prod:
            vid    = v.get("id")
            color  = normalizar_color(v.get("color") or "")
            talla  = str(v.get("talla") or "").strip()
            sku    = (v.get("sku") or "").strip()
            stock  = stock_map.get(vid, 1)  # si no hay inventario, poner 1
            foto   = v.get("foto_url") or ""
            # Si no hay foto_url, usar la primera de imagenes[]
            if not foto and v.get("imagenes"):
                imgs = v["imagenes"]
                if isinstance(imgs, list) and imgs:
                    foto = imgs[0]
                elif isinstance(imgs, str):
                    try:
                        lst = json.loads(imgs)
                        foto = lst[0] if lst else ""
                    except Exception:
                        foto = imgs

            if sheet == "Zapatillas y Tacones":
                # 33 columnas
                fila = [
                    titulo,           # 1  Título
                    len(titulo),      # 2  Cantidad caracteres
                    CONDICION,        # 3  Condición
                    color,            # 4  Color
                    "Liso",           # 5  Diseño de la tela
                    talla,            # 6  Talla
                    MARCA,            # 7  Marca
                    genero,           # 8  Género
                    guia,             # 9  Código de la guía
                    foto,             # 10 Fotos
                    COD_UNIV,         # 11 Código universal
                    sku,              # 12 SKU
                    stock,            # 13 Stock
                    precio,           # 14 Precio
                    MONEDA,           # 15 Moneda
                    desc,             # 16 Descripción
                    TIPO_PUB,         # 17 Tipo de publicación
                    None,             # 18 Cargo por venta
                    FORMA_ENVIO,      # 19 Forma de envío
                    COSTO_ENVIO,      # 20 Costo de envío
                    RETIRO,           # 21 Retiro en persona
                    None,             # 22 Tipo de garantía
                    None,             # 23 Tiempo de garantía
                    None,             # 24 Unidad de tiempo garantía
                    modelo,           # 25 Modelo
                    tipo_cal,         # 26 Tipo de calzado
                    mat_cal,          # 27 Material del calzado
                    mat_suel,         # 28 Material de la suela
                    tipo_tac,         # 29 Tipo de taco
                    None,             # 30 Temporada
                    None,             # 31 Año
                    None,             # 32 BUYBOX
                    None,             # 33 HIDDEN_PICTURES
                ]

            elif sheet == "Sandalias y Chanclas":
                # 34 columnas (col 1 = Código catálogo ML — vacío para nuevas)
                fila = [
                    None,             # 1  Código catálogo ML
                    titulo,           # 2  Título
                    len(titulo),      # 3  Cantidad caracteres
                    CONDICION,        # 4  Condición
                    color,            # 5  Color
                    "Liso",           # 6  Nombre del diseño
                    talla,            # 7  Talla
                    MARCA,            # 8  Marca
                    genero,           # 9  Género
                    guia,             # 10 Código de la guía
                    foto,             # 11 Fotos
                    COD_UNIV,         # 12 Código universal
                    sku,              # 13 SKU
                    stock,            # 14 Stock
                    precio,           # 15 Precio
                    MONEDA,           # 16 Moneda
                    desc,             # 17 Descripción
                    TIPO_PUB,         # 18 Tipo de publicación
                    None,             # 19 Cargo por venta
                    FORMA_ENVIO,      # 20 Forma de envío
                    COSTO_ENVIO,      # 21 Costo de envío
                    RETIRO,           # 22 Retiro en persona
                    None,             # 23 Tipo de garantía
                    None,             # 24 Tiempo de garantía
                    None,             # 25 Unidad de tiempo garantía
                    modelo,           # 26 Modelo
                    "Sandalias altas" if (prod.get("altura_tacon") or 0) > 3 else "Escribe o elige un valor",  # 27 Tipo calzado
                    None,             # 28 Estilo del calzado
                    mat_cal,          # 29 Materiales del calzado
                    mat_suel,         # 30 Material de la suela
                    None,             # 31 Temporada
                    None,             # 32 Año
                    None,             # 33 BUYBOX
                    None,             # 34 HIDDEN_PICTURES
                ]

            elif sheet == "Mocasines y Oxfords":
                # 38 columnas
                fila = [
                    None,             # 1  Código catálogo ML
                    titulo,           # 2  Título
                    len(titulo),      # 3  Cantidad caracteres
                    CONDICION,        # 4  Condición
                    color,            # 5  Color
                    "Liso",           # 6  Diseño de la tela
                    talla,            # 7  Talla
                    MARCA,            # 8  Marca
                    genero,           # 9  Género
                    guia,             # 10 Código de la guía
                    foto,             # 11 Fotos
                    COD_UNIV,         # 12 Código universal
                    sku,              # 13 SKU
                    stock,            # 14 Stock
                    precio,           # 15 Precio
                    MONEDA,           # 16 Moneda
                    desc,             # 17 Descripción
                    TIPO_PUB,         # 18 Tipo de publicación
                    None,             # 19 Cargo por venta
                    FORMA_ENVIO,      # 20 Forma de envío
                    COSTO_ENVIO,      # 21 Costo de envío
                    RETIRO,           # 22 Retiro en persona
                    None,             # 23 Tipo de garantía
                    None,             # 24 Tiempo de garantía
                    None,             # 25 Unidad de tiempo garantía
                    modelo,           # 26 Modelo
                    None,             # 27 Tipo de calzado
                    None,             # 28 Estilo del calzado
                    None,             # 29 Tipo de perforación
                    None,             # 30 Tipo de textura
                    None,             # 31 Forma de la punta
                    None,             # 32 Tipo de ajuste
                    mat_cal,          # 33 Material externo
                    None,             # 34 Material interno
                    mat_suel,         # 35 Material de la suela
                    None,             # 36 Talón plegable
                    None,             # 37 BUYBOX
                    None,             # 38 HIDDEN_PICTURES
                ]
            else:
                continue

            sheets[sheet].append(fila)

    return sheets

# ─── Escritura del Excel ──────────────────────────────────────────────────────

def _safe_set(ws, row, col, value):
    """Escribe en una celda ignorando MergedCell (solo-lectura)."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value

def escribir_excel(sheets_data, template_path, output_path):
    """Toma el template de ML como base, limpia las filas de datos y escribe las nuestras."""
    print(f"\n[ ] Cargando template: {template_path}")
    wb = openpyxl.load_workbook(template_path)

    for sheet_name, filas in sheets_data.items():
        if not filas:
            print(f"   !  {sheet_name}: sin datos, se omite")
            continue

        if sheet_name not in wb.sheetnames:
            print(f"   !  Hoja '{sheet_name}' no encontrada en el template")
            continue

        ws = wb[sheet_name]

        # Deshacer combinaciones de celdas en el area de datos (fila 8+)
        merged_to_remove = [
            r for r in list(ws.merged_cells.ranges)
            if r.min_row >= 8
        ]
        for r in merged_to_remove:
            ws.unmerge_cells(str(r))

        # Limpiar filas de datos existentes (fila 8 en adelante)
        max_row = ws.max_row
        for row_idx in range(8, max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                _safe_set(ws, row_idx, col_idx, None)

        # Escribir nuestras filas
        for i, fila in enumerate(filas):
            row_idx = 8 + i
            for col_idx, valor in enumerate(fila, 1):
                _safe_set(ws, row_idx, col_idx, valor)

        print(f"   OK  {sheet_name}: {len(filas)} variantes escritas")

    wb.save(output_path)
    print(f"\nLISTO Archivo guardado en:\n   {output_path}")

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    # Encontrar el template
    template_path = TEMPLATE_DEFAULT
    if len(sys.argv) > 1:
        template_path = sys.argv[1]

    if not os.path.exists(template_path):
        print(f"ERROR: No encuentro el template en:\n  {template_path}")
        print("Uso: python exportar_mercadolibre.py [ruta_al_template.xlsx]")
        sys.exit(1)

    # Nombre del archivo de salida
    fecha = datetime.now().strftime("%Y-%m-%d_%H%M")
    output_dir = os.path.expanduser("~/Desktop")
    output_path = os.path.join(output_dir, f"ML_ZapatillasMay_{fecha}.xlsx")

    print("=" * 60)
    print("  ERP Zapatillas May -> Exportador MercadoLibre")
    print("=" * 60)

    # Cargar datos
    prod_map, variantes, stock_map = cargar_datos()

    # Construir filas
    print("\n[ ] Generando filas…")
    sheets_data = construir_filas(prod_map, variantes, stock_map)

    totales = {k: len(v) for k, v in sheets_data.items()}
    print(f"   Zapatillas y Tacones : {totales['Zapatillas y Tacones']:>5} variantes")
    print(f"   Sandalias y Chanclas : {totales['Sandalias y Chanclas']:>5} variantes")
    print(f"   Mocasines y Oxfords  : {totales['Mocasines y Oxfords']:>5} variantes")
    print(f"   TOTAL                : {sum(totales.values()):>5} variantes")

    # Escribir Excel
    escribir_excel(sheets_data, template_path, output_path)

    print("\nRECUERDA antes de subir a ML:")
    print("   1. Verifica que el código de guía sea el correcto para cada producto")
    print("   2. Las fotos son URLs de Cloudinary — súbelas al Gestor de fotos de ML")
    print("      y reemplaza las URLs en la columna 'Fotos'")
    print("   3. Revisa los títulos (máx 60 caracteres)")
    print("   4. Ajusta los precios si necesitas incluir el costo de envío")
    print("=" * 60)

if __name__ == "__main__":
    main()
